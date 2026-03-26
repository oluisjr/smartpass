from openpyxl import load_workbook
from io import BytesIO

def read_excel(buffer: bytes):
    wb = load_workbook(filename=BytesIO(buffer))
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    return rows
